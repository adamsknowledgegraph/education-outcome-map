from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


ROOT = Path("/Users/adam/Agents/degree-career-map")
DATA_DIR = ROOT / "data"
RAW_ONET_DIR = ROOT / "raw" / "onet" / "db_30_2_text" / "db_30_2_text"
CIP_CROSSWALK = ROOT / "raw" / "crosswalks" / "Education_CIP_to_ONET_SOC.xlsx"
APP_DIR = ROOT / "app"
APP_DATA_JS = APP_DIR / "data" / "prototype-data.js"
APP_GRAPH_JSON = APP_DIR / "data" / "grc20-compatible-graph.json"

INTEREST_META = {
    "tech": {
        "interest_id": "tech",
        "interest_name": "Technology",
        "description": "Building software, systems, tools, and digital products.",
    },
    "healthcare": {
        "interest_id": "healthcare",
        "interest_name": "Healthcare",
        "description": "Helping people through medical, clinical, or care work.",
    },
    "design": {
        "interest_id": "design",
        "interest_name": "Design",
        "description": "Visual, digital, and product experience work.",
    },
    "business": {
        "interest_id": "business",
        "interest_name": "Business",
        "description": "Strategy, finance, operations, commercial growth, and management.",
    },
    "research": {
        "interest_id": "research",
        "interest_name": "Research",
        "description": "Analysis, evidence, investigation, and structured problem-solving.",
    },
    "communication": {
        "interest_id": "communication",
        "interest_name": "Communication",
        "description": "Writing, presenting, teaching, persuading, or coordinating.",
    },
    "hands_on": {
        "interest_id": "hands_on",
        "interest_name": "Hands-on Work",
        "description": "Practical work with physical systems, tools, or visible output.",
    },
    "helping_people": {
        "interest_id": "helping_people",
        "interest_name": "Helping People",
        "description": "Service, support, care, guidance, and impact-oriented work.",
    },
}

ACADEMIC_PROFILE_ORDER = {
    "open": 1,
    "accessible": 2,
    "solid": 3,
    "strong": 4,
    "elite": 5,
}

COST_BAND_ORDER = {
    "low": 1,
    "medium": 2,
    "high": 3,
}

SOURCE_NOTES = {
    "onet_30_2": "Skills, knowledge, interests, and work-context values come from the O*NET 30.2 database.",
    "cip_crosswalk": "Education-to-career links use the official 2020 CIP to O*NET-SOC crosswalk where available.",
    "bls_ooh_2026": "Occupation wages, outlook, and quick facts come from BLS Occupational Outlook Handbook pages accessed on March 12, 2026.",
    "bls_field_of_degree_2025": "Field-of-degree context comes from official BLS field-of-degree pages.",
    "manual": "Manual route included to cover real-world pathways not represented cleanly in public crosswalks.",
    "manual_plus_cip": "Path combines manual curation with the official CIP-to-O*NET crosswalk.",
    "manual_plus_related": "Path is a plausible adjacent route rather than the default or most direct outcome.",
}


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def load_csv_optional(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    return load_csv(path)


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_json(path: Path, payload: dict | list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_cip_crosswalk(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    for cip_code, cip_title, onet_soc_code, onet_title in sheet.iter_rows(min_row=5, values_only=True):
        if not cip_code:
            continue
        rows.append(
            {
                "cip_code": str(cip_code),
                "cip_title": str(cip_title),
                "onet_soc_code": str(onet_soc_code),
                "onet_title": str(onet_title),
            }
        )
    return rows


def slug(text: str) -> str:
    return (
        text.lower()
        .replace("&", "and")
        .replace("/", " ")
        .replace(",", "")
        .replace(".", "")
        .replace("-", " ")
        .replace("(", "")
        .replace(")", "")
        .replace("'", "")
    ).strip().replace(" ", "_")


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def to_float(value: str | int | float | None) -> float:
    if value in ("", None):
        return 0.0
    return float(value)


def to_int(value: str | int | float | None) -> int:
    if value in ("", None):
        return 0
    return int(float(value))


def split_tags(value: str) -> list[str]:
    return [item.strip() for item in value.split(";") if item.strip()]


def title_case(value: str) -> str:
    return value.replace("_", " ").title()


def onet_online_url(code: str) -> str:
    return f"https://www.onetonline.org/link/summary/{code}"


def load_ranked_measure(rows: list[dict[str, str]], scale_id: str) -> dict[str, list[dict[str, float | str]]]:
    grouped: dict[str, list[dict[str, float | str]]] = defaultdict(list)
    for row in rows:
        if row.get("Scale ID") != scale_id:
            continue
        grouped[row["O*NET-SOC Code"]].append(
            {
                "name": row["Element Name"],
                "importance_score": round(float(row["Data Value"]), 2),
                "level_score": 0.0,
            }
        )
    for code, items in grouped.items():
        deduped = {}
        for item in items:
            deduped[item["name"]] = item
        grouped[code] = sorted(deduped.values(), key=lambda item: item["importance_score"], reverse=True)
    return grouped


def load_level_lookup(rows: list[dict[str, str]], scale_id: str) -> dict[tuple[str, str], float]:
    lookup = {}
    for row in rows:
        if row.get("Scale ID") != scale_id:
            continue
        lookup[(row["O*NET-SOC Code"], row["Element Name"])] = round(float(row["Data Value"]), 2)
    return lookup


def load_category_descriptions(rows: list[dict[str, str]]) -> dict[tuple[str, str, str], str]:
    lookup: dict[tuple[str, str, str], str] = {}
    for row in rows:
        lookup[(row["Element Name"], row["Scale ID"], row["Category"])] = row["Category Description"]
    return lookup


def load_category_distribution(
    rows: list[dict[str, str]],
    element_name: str,
    scale_id: str,
    category_lookup: dict[tuple[str, str, str], str],
) -> dict[str, list[dict[str, float | str]]]:
    grouped: dict[str, list[dict[str, float | str]]] = defaultdict(list)
    for row in rows:
        if row.get("Element Name") != element_name or row.get("Scale ID") != scale_id:
            continue
        grouped[row["O*NET-SOC Code"]].append(
            {
                "category": row["Category"],
                "label": category_lookup.get((element_name, scale_id, row["Category"]), row["Category"]),
                "value": round(float(row["Data Value"]), 2),
            }
        )
    for code in grouped:
        grouped[code] = sorted(grouped[code], key=lambda item: item["value"], reverse=True)
    return grouped


def top_work_styles(rows: list[dict[str, str]]) -> dict[str, list[dict[str, float | str]]]:
    grouped: dict[str, list[dict[str, float | str]]] = defaultdict(list)
    for row in rows:
        if row.get("Scale ID") != "WI":
            continue
        grouped[row["O*NET-SOC Code"]].append(
            {
                "name": row["Element Name"],
                "score": round(float(row["Data Value"]), 2),
            }
        )
    for code in grouped:
        grouped[code] = sorted(grouped[code], key=lambda item: item["score"], reverse=True)
    return grouped


def technology_examples(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    seen: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        code = row["O*NET-SOC Code"]
        example = row["Example"]
        if example in seen[code]:
            continue
        seen[code].add(example)
        grouped[code].append(
            {
                "example": example,
                "commodity_title": row["Commodity Title"],
                "hot_technology": row["Hot Technology"],
                "in_demand": row["In Demand"],
            }
        )
    for code in grouped:
        grouped[code] = sorted(
            grouped[code],
            key=lambda item: (item["hot_technology"] == "Y", item["in_demand"] == "Y", item["example"]),
            reverse=True,
        )
    return grouped


def related_occupations(rows: list[dict[str, str]], occupation_lookup: dict[str, dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        related_code = row["Related O*NET-SOC Code"]
        related_title = occupation_lookup.get(related_code, {}).get("Title")
        if not related_title:
            continue
        grouped[row["O*NET-SOC Code"]].append(
            {
                "soc_code": related_code,
                "title": related_title,
                "tier": row["Relatedness Tier"],
            }
        )
    return grouped


def top_interest_profile(rows: list[dict[str, str]]) -> dict[str, list[dict[str, float | str]]]:
    grouped: dict[str, list[dict[str, float | str]]] = defaultdict(list)
    for row in rows:
        grouped[row["O*NET-SOC Code"]].append(
            {"name": row["Element Name"], "score": round(float(row["Data Value"]), 2)}
        )
    for code in grouped:
        grouped[code] = sorted(grouped[code], key=lambda item: item["score"], reverse=True)
    return grouped


def work_context_lookup(rows: list[dict[str, str]]) -> dict[str, dict[str, float]]:
    grouped: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        code = row["O*NET-SOC Code"]
        element = row["Element Name"]
        scale = row["Scale ID"]
        value = float(row["Data Value"])
        if scale == "CX":
            grouped[code][element].append(value)
        elif scale in {"CXP", "CTP"}:
            grouped[code][f"{element}__{row['Category']}"].append(value)

    flattened: dict[str, dict[str, float]] = defaultdict(dict)
    for code, items in grouped.items():
        for key, values in items.items():
            flattened[code][key] = round(mean(values), 2)
    return flattened


def derive_schedule_descriptor(context: dict[str, float], override: str) -> str:
    if override:
        return override
    regular = context.get("Work Schedules__1", 0.0)
    irregular = context.get("Work Schedules__2", 0.0)
    seasonal = context.get("Work Schedules__3", 0.0)
    if irregular >= max(regular, seasonal):
        return "irregular"
    if seasonal > regular:
        return "seasonal"
    return "normal"


def source_link(
    source_id: str,
    label: str,
    url: str,
    source_type: str,
    note: str,
    publisher: str = "",
) -> dict[str, str]:
    return {
        "source_id": source_id,
        "label": label,
        "url": url,
        "source_type": source_type,
        "publisher": publisher,
        "note": note,
    }


def dedupe_sources(items: list[dict[str, str]]) -> list[dict[str, str]]:
    ordered: list[dict[str, str]] = []
    seen = set()
    for item in items:
        key = item["source_id"]
        if key in seen:
            continue
        ordered.append(item)
        seen.add(key)
    return ordered


def strength_fit_bundle(row: dict[str, str]) -> dict[str, int]:
    return {
        "analytical": to_int(row.get("analytical_fit")),
        "creative": to_int(row.get("creative_fit")),
        "people": to_int(row.get("people_fit")),
        "hands_on": to_int(row.get("hands_on_fit")),
    }


def build_tradeoff_summary(path: dict[str, str | int | float], target_income: float) -> str:
    annual = float(path["median_annual_wage_usd"])
    hours = int(path["avg_weekly_hours_estimate"])
    ai = path["ai_exposure"]
    geo = path["geographic_concentration"]
    schedule = path["schedule_type"]
    if annual >= target_income and hours >= 48:
        return "Can fund the target lifestyle, but it usually asks for longer weeks or weaker boundaries."
    if annual >= target_income and geo == "major_hub_dependent":
        return "The earnings can work, but the upside often pulls you toward major city hubs."
    if ai == "low" and geo == "local_anywhere":
        return "More geographically flexible and less exposed to AI, with a grounded day-to-day tradeoff."
    if schedule in {"shift", "irregular"}:
        return "The money may work, but your calendar is likely to be the real compromise."
    return "A more balanced path where the tradeoffs are easier to explain than annual salary alone."


def build_reality_check(degree: dict[str, str], profession: dict[str, str], path: dict[str, str]) -> str:
    years = to_int(path["estimated_total_training_years"])
    confidence = to_float(path["confidence"])
    extra_training = path["additional_training_required"]
    if extra_training == "medical_school" or years >= 8:
        return "Possible, but this is a long and highly selective training path rather than a standard default outcome."
    if confidence < 0.5:
        return "Treat this as an adjacent possibility, not the most common outcome from this education route."
    if degree["pathway_type"] == "trade":
        return "This route trades prestige signaling for faster entry, lower debt, and more location flexibility."
    if profession["ai_exposure"] == "high":
        return "This path may stay viable, but students should expect a lot of the routine work to be reshaped by AI."
    if profession["schedule_type"] in {"shift", "irregular"}:
        return "The real cost here is schedule control, not just training time."
    return "The route is reasonably direct, but the best outcomes still depend on execution, internships, or licensure."


def lifestyle_band(hours: int) -> str:
    if hours >= 60:
        return "60_plus"
    if hours >= 50:
        return "50_60"
    if hours >= 40:
        return "40_50"
    return "35_45"


def json_compact(value: object) -> str:
    return json.dumps(value, ensure_ascii=True)


def main() -> None:
    manual_degrees = load_csv(DATA_DIR / "manual_degrees.csv")
    manual_professions = load_csv(DATA_DIR / "manual_profession_overrides.csv")
    manual_paths = load_csv(DATA_DIR / "manual_path_overrides.csv")
    lifestyle_tiers = load_csv(DATA_DIR / "lifestyle_tiers.csv")
    city_cost_tiers = load_csv(DATA_DIR / "city_cost_tiers.csv")
    education_programs = load_csv_optional(DATA_DIR / "education_programs_v1.csv")
    cip_crosswalk = load_cip_crosswalk(CIP_CROSSWALK)

    occupation_rows = load_tsv(RAW_ONET_DIR / "Occupation Data.txt")
    skills_rows = load_tsv(RAW_ONET_DIR / "Skills.txt")
    knowledge_rows = load_tsv(RAW_ONET_DIR / "Knowledge.txt")
    interests_rows = load_tsv(RAW_ONET_DIR / "Interests.txt")
    work_context_rows = load_tsv(RAW_ONET_DIR / "Work Context.txt")
    job_zone_rows = load_tsv(RAW_ONET_DIR / "Job Zones.txt")
    education_training_rows = load_tsv(RAW_ONET_DIR / "Education, Training, and Experience.txt")
    education_training_categories = load_tsv(
        RAW_ONET_DIR / "Education, Training, and Experience Categories.txt"
    )
    work_styles_rows = load_tsv(RAW_ONET_DIR / "Work Styles.txt")
    technology_rows = load_tsv(RAW_ONET_DIR / "Technology Skills.txt")
    related_occupations_rows = load_tsv(RAW_ONET_DIR / "Related Occupations.txt")

    occupation_lookup = {row["O*NET-SOC Code"]: row for row in occupation_rows}
    skill_importance = load_ranked_measure(skills_rows, "IM")
    skill_levels = load_level_lookup(skills_rows, "LV")
    knowledge_importance = load_ranked_measure(knowledge_rows, "IM")
    knowledge_levels = load_level_lookup(knowledge_rows, "LV")
    interest_profiles = top_interest_profile(interests_rows)
    context_lookup = work_context_lookup(work_context_rows)
    job_zone_lookup = {row["O*NET-SOC Code"]: row["Job Zone"] for row in job_zone_rows}
    category_lookup = load_category_descriptions(education_training_categories)
    required_education_distribution = load_category_distribution(
        education_training_rows,
        "Required Level of Education",
        "RL",
        category_lookup,
    )
    related_experience_distribution = load_category_distribution(
        education_training_rows,
        "Related Work Experience",
        "RW",
        category_lookup,
    )
    on_job_training_distribution = load_category_distribution(
        education_training_rows,
        "On-the-Job Training",
        "OJ",
        category_lookup,
    )
    work_styles_lookup = top_work_styles(work_styles_rows)
    technology_lookup = technology_examples(technology_rows)
    related_occupations_lookup = related_occupations(related_occupations_rows, occupation_lookup)

    crosswalk_by_cip: dict[str, set[str]] = defaultdict(set)
    crosswalk_titles: dict[tuple[str, str], str] = {}
    for row in cip_crosswalk:
        crosswalk_by_cip[row["cip_code"]].add(row["onet_soc_code"])
        crosswalk_titles[(row["cip_code"], row["onet_soc_code"])] = row["onet_title"]

    source_registry: dict[str, dict[str, str]] = {}
    source_registry["source:onet:release"] = source_link(
        "source:onet:release",
        "O*NET 30.2 database",
        "https://www.onetcenter.org/database.html",
        "dataset",
        SOURCE_NOTES["onet_30_2"],
        "O*NET / U.S. Department of Labor",
    )
    source_registry["source:cip:crosswalk"] = source_link(
        "source:cip:crosswalk",
        "2020 CIP to O*NET-SOC crosswalk",
        "https://www.onetcenter.org/crosswalks.html",
        "dataset",
        SOURCE_NOTES["cip_crosswalk"],
        "O*NET / U.S. Department of Labor",
    )
    source_registry["source:overlay:heuristics"] = source_link(
        "source:overlay:heuristics",
        "Student-facing lifestyle overlay",
        "",
        "heuristic_overlay",
        "Boundary quality, AI exposure, and geographic concentration combine O*NET work-context data with manual interpretation to stay useful for students.",
        "Education Outcome Map prototype",
    )
    source_registry["source:scorecard:dataset"] = source_link(
        "source:scorecard:dataset",
        "College Scorecard institution and field-of-study data",
        "https://collegescorecard.ed.gov/data/",
        "dataset",
        "Institution and program examples come from the U.S. Department of Education College Scorecard data layer.",
        "U.S. Department of Education",
    )

    degrees_out: list[dict[str, object]] = []
    degree_interest_tags: list[dict[str, object]] = []
    degree_records: dict[str, dict[str, object]] = {}

    for degree in manual_degrees:
        degree_id = degree["degree_id"]
        node_id = f"degree:{degree_id}"
        field_source_id = f"source:degree-field:{degree_id}"
        degree_sources = []
        if degree["field_of_degree_url"]:
            source_registry[field_source_id] = source_link(
                field_source_id,
                f"{degree['degree_name']} field evidence",
                degree["field_of_degree_url"],
                "field_of_degree",
                SOURCE_NOTES.get(degree["field_source"], ""),
                "BLS",
            )
            degree_sources.append(source_registry[field_source_id])
        if degree["cip_code"]:
            degree_sources.append(source_registry["source:cip:crosswalk"])

        degree_record = {
            "degree_id": degree_id,
            "degree_node_id": node_id,
            "degree_name": degree["degree_name"],
            "degree_family": degree["degree_family"],
            "cip_code": degree["cip_code"],
            "award_level": degree["award_level"],
            "pathway_type": degree["pathway_type"],
            "source_type": degree["source_type"],
            "notes": degree["notes"],
            "academic_profile_min": degree["academic_profile_min"],
            "academic_profile_min_rank": ACADEMIC_PROFILE_ORDER[degree["academic_profile_min"]],
            "time_to_complete_years": to_int(degree["time_to_complete_years"]),
            "cost_band": degree["cost_band"],
            "cost_band_rank": COST_BAND_ORDER[degree["cost_band"]],
            "fit_scores": strength_fit_bundle(degree),
            "field_of_degree_url": degree["field_of_degree_url"],
            "field_employment_2023": to_int(degree["field_employment_2023"]),
            "field_median_annual_wage_usd": to_int(degree["field_median_annual_wage_usd"]),
            "field_pct_bachelors_jobs": to_int(degree["field_pct_bachelors_jobs"]),
            "field_pct_advanced_degree": to_int(degree["field_pct_advanced_degree"]),
            "field_source": degree["field_source"],
            "grade_proxy_note": degree["grade_proxy_note"],
            "interest_tags": split_tags(degree["interest_tags"]),
            "source_links": dedupe_sources(degree_sources),
        }
        degree_records[degree_id] = degree_record
        degrees_out.append(degree_record)
        for interest in degree_record["interest_tags"]:
            degree_interest_tags.append(
                {
                    "degree_id": degree_id,
                    "interest_id": interest,
                    "edge_type": "tagged_with",
                    "weight": 1.0,
                }
            )

    professions_out: list[dict[str, object]] = []
    lifestyle_out: list[dict[str, object]] = []
    profession_interest_tags: list[dict[str, object]] = []
    profession_skill_rows: list[dict[str, object]] = []
    unique_skills: dict[str, dict[str, object]] = {}
    profession_records: dict[str, dict[str, object]] = {}
    profession_by_soc: dict[str, dict[str, object]] = {}

    for profession in manual_professions:
        code = profession["onet_soc_code"]
        occupation = occupation_lookup.get(code, {})
        context = context_lookup.get(code, {})
        education_distribution = required_education_distribution.get(code, [])[:4]
        experience_distribution = related_experience_distribution.get(code, [])[:3]
        training_distribution = on_job_training_distribution.get(code, [])[:3]
        work_styles = work_styles_lookup.get(code, [])[:4]
        tech_examples = technology_lookup.get(code, [])[:6]
        related_roles = related_occupations_lookup.get(code, [])[:5]
        top_skills = skill_importance.get(code, [])[:5]
        for item in top_skills:
            item["level_score"] = skill_levels.get((code, item["name"]), 0.0)
        top_knowledge = knowledge_importance.get(code, [])[:3]
        for item in top_knowledge:
            item["level_score"] = knowledge_levels.get((code, item["name"]), 0.0)

        for item in top_skills + top_knowledge:
            skill_id = slug(item["name"])
            unique_skills[skill_id] = {
                "skill_id": skill_id,
                "skill_node_id": f"skill:{skill_id}",
                "skill_name": item["name"],
                "skill_category": "knowledge" if item in top_knowledge else "skill",
                "source": "onet_30_2",
            }
            profession_skill_rows.append(
                {
                    "profession_id": profession["profession_id"],
                    "skill_id": skill_id,
                    "edge_type": "requires",
                    "importance_score": round(float(item["importance_score"]), 2),
                    "level_score": round(float(item["level_score"]), 2),
                    "source": "onet_30_2",
                }
            )

        for interest in split_tags(profession["interest_tags"]):
            profession_interest_tags.append(
                {
                    "profession_id": profession["profession_id"],
                    "interest_id": interest,
                    "edge_type": "tagged_with",
                    "weight": 1.0,
                }
            )

        ooh_source_id = f"source:ooh:{profession['profession_id']}"
        onet_source_id = f"source:onet:{profession['profession_id']}"
        source_registry[ooh_source_id] = source_link(
            ooh_source_id,
            f"{profession['profession_name']} OOH page",
            profession["ooh_url"],
            "occupation_outlook",
            SOURCE_NOTES.get(profession["source"], SOURCE_NOTES["bls_ooh_2026"]),
            "BLS",
        )
        source_registry[onet_source_id] = source_link(
            onet_source_id,
            f"{profession['profession_name']} O*NET profile",
            onet_online_url(code),
            "occupation_profile",
            SOURCE_NOTES["onet_30_2"],
            "O*NET / U.S. Department of Labor",
        )

        lifestyle_record = {
            "profession_id": profession["profession_id"],
            "lifestyle_profile_node_id": f"lifestyle:{profession['profession_id']}",
            "avg_weekly_hours_band": lifestyle_band(to_int(profession["avg_weekly_hours_estimate"])),
            "avg_weekly_hours_estimate": to_int(profession["avg_weekly_hours_estimate"]),
            "schedule_type": derive_schedule_descriptor(context, profession["schedule_type"]),
            "boundary_quality": profession["boundary_quality"],
            "predictability": profession["predictability"],
            "ai_exposure": profession["ai_exposure"],
            "geographic_concentration": profession["geographic_concentration"],
            "physical_intensity": profession["physical_intensity"],
            "remote_friendliness": profession["remote_friendliness"],
            "weekend_work_likelihood": profession["weekend_work_likelihood"],
            "night_work_likelihood": profession["night_work_likelihood"],
            "heuristic_flag": "mixed",
            "source": f"{profession['source']};onet_30_2",
            "notes": f"{profession['hours_note']} {profession['notes']}".strip(),
            "time_pressure_score": round(context.get("Time Pressure", 0.0), 2),
            "contact_with_others_score": round(context.get("Contact With Others", 0.0), 2),
            "email_score": round(context.get("E-Mail", 0.0), 2),
            "face_to_face_score": round(
                context.get("Face-to-Face Discussions with Individuals and Within Teams", 0.0), 2
            ),
            "sitting_score": round(context.get("Spend Time Sitting", 0.0), 2),
            "standing_score": round(context.get("Spend Time Standing", 0.0), 2),
            "walking_score": round(context.get("Spend Time Walking or Running", 0.0), 2),
            "regular_schedule_pct": round(context.get("Work Schedules__1", 0.0), 2),
            "irregular_schedule_pct": round(context.get("Work Schedules__2", 0.0), 2),
            "seasonal_schedule_pct": round(context.get("Work Schedules__3", 0.0), 2),
            "required_education_distribution": education_distribution,
            "related_experience_distribution": experience_distribution,
            "on_job_training_distribution": training_distribution,
        }
        lifestyle_out.append(lifestyle_record)

        ranked_skills = sorted(
            [row for row in profession_skill_rows if row["profession_id"] == profession["profession_id"]],
            key=lambda item: (item["importance_score"], item["level_score"]),
            reverse=True,
        )
        top_skill_bundle = []
        for skill_row in ranked_skills[:5]:
            skill = unique_skills[skill_row["skill_id"]]
            top_skill_bundle.append(
                {
                    "skill_id": skill["skill_id"],
                    "skill_name": skill["skill_name"],
                    "skill_category": skill["skill_category"],
                    "importance_score": skill_row["importance_score"],
                    "level_score": skill_row["level_score"],
                }
            )

        profession_record = {
            "profession_id": profession["profession_id"],
            "profession_node_id": f"profession:{profession['profession_id']}",
            "profession_name": profession["profession_name"],
            "soc_code": code,
            "profession_family": profession["profession_family"],
            "median_annual_wage_usd": to_int(profession["median_annual_wage_usd"]),
            "median_hourly_wage_usd": round(float(profession["median_hourly_wage_usd"]), 2),
            "employment_count": to_int(profession["employment_count"]),
            "outlook_growth_pct": to_int(profession["outlook_growth_pct"]),
            "openings_annual": to_int(profession["openings_annual"]),
            "job_zone": job_zone_lookup.get(code, ""),
            "onet_title": occupation.get("Title", profession["profession_name"]),
            "onet_description": occupation.get("Description", ""),
            "entry_level_education": profession["entry_level_education"],
            "typical_training_years_post_hs": to_int(profession["typical_training_years_post_hs"]),
            "hours_note": profession["hours_note"],
            "notes": profession["notes"],
            "interest_tags": split_tags(profession["interest_tags"]),
            "source": profession["source"],
            "ooh_url": profession["ooh_url"],
            "onet_url": onet_online_url(code),
            "source_links": dedupe_sources(
                [
                    source_registry[ooh_source_id],
                    source_registry[onet_source_id],
                    source_registry["source:overlay:heuristics"],
                ]
            ),
            "top_skills": top_skill_bundle,
            "top_work_styles": work_styles,
            "technology_examples": tech_examples,
            "related_roles": related_roles,
            "required_education_distribution": education_distribution,
            "related_experience_distribution": experience_distribution,
            "on_job_training_distribution": training_distribution,
            "interest_profile": interest_profiles.get(code, [])[:3],
            **lifestyle_record,
        }
        profession_records[profession["profession_id"]] = profession_record
        profession_by_soc[code] = profession_record
        professions_out.append(profession_record)

    degree_to_profession_out: list[dict[str, object]] = []
    path_outcomes: list[dict[str, object]] = []

    default_target_income = 85000.0
    for path in manual_paths:
        degree = degree_records[path["degree_id"]]
        profession = profession_by_soc[path["onet_soc_code"]]
        crosswalk_match = False
        if degree["cip_code"]:
            crosswalk_match = path["onet_soc_code"] in crosswalk_by_cip.get(str(degree["cip_code"]), set())

        relation_sources = list(degree["source_links"]) + list(profession["source_links"])
        if degree["cip_code"] and crosswalk_match:
            relation_sources.append(source_registry["source:cip:crosswalk"])
        relation_sources.append(source_registry["source:overlay:heuristics"])
        relation_sources = dedupe_sources(relation_sources)
        source_ids = [item["source_id"] for item in relation_sources]

        degree_to_profession_out.append(
            {
                "degree_id": degree["degree_id"],
                "profession_id": profession["profession_id"],
                "edge_type": "leads_to",
                "path_strength": path["path_strength"],
                "confidence": round(to_float(path["confidence"]), 2),
                "source": path["source"],
                "notes": path["path_note"],
                "crosswalk_match": "yes" if crosswalk_match else "no",
                "additional_training_required": path["additional_training_required"],
                "estimated_total_training_years": to_int(path["estimated_total_training_years"]),
                "source_ids_json": json_compact(source_ids),
            }
        )

        interest_matches = sorted(
            set(degree["interest_tags"]) & set(profession["interest_tags"])
        )
        top_skills = profession["top_skills"][:3]
        path_record = {
            "path_id": f"path:{degree['degree_id']}:{profession['profession_id']}",
            "degree_id": degree["degree_id"],
            "profession_id": profession["profession_id"],
            "degree_name": degree["degree_name"],
            "profession_name": profession["profession_name"],
            "degree_family": degree["degree_family"],
            "profession_family": profession["profession_family"],
            "soc_code": profession["soc_code"],
            "award_level": degree["award_level"],
            "pathway_type": degree["pathway_type"],
            "time_to_complete_years": degree["time_to_complete_years"],
            "estimated_total_training_years": to_int(path["estimated_total_training_years"]),
            "academic_profile_min": degree["academic_profile_min"],
            "cost_band": degree["cost_band"],
            "entry_level_education": profession["entry_level_education"],
            "typical_training_years_post_hs": profession["typical_training_years_post_hs"],
            "additional_training_required": path["additional_training_required"],
            "path_strength": path["path_strength"],
            "confidence": round(to_float(path["confidence"]), 2),
            "crosswalk_match": crosswalk_match,
            "crosswalk_onet_title": crosswalk_titles.get((str(degree["cip_code"]), profession["soc_code"]), ""),
            "median_annual_wage_usd": profession["median_annual_wage_usd"],
            "median_hourly_wage_usd": profession["median_hourly_wage_usd"],
            "employment_count": profession["employment_count"],
            "outlook_growth_pct": profession["outlook_growth_pct"],
            "openings_annual": profession["openings_annual"],
            "avg_weekly_hours_estimate": profession["avg_weekly_hours_estimate"],
            "schedule_type": profession["schedule_type"],
            "boundary_quality": profession["boundary_quality"],
            "predictability": profession["predictability"],
            "ai_exposure": profession["ai_exposure"],
            "geographic_concentration": profession["geographic_concentration"],
            "physical_intensity": profession["physical_intensity"],
            "remote_friendliness": profession["remote_friendliness"],
            "weekend_work_likelihood": profession["weekend_work_likelihood"],
            "night_work_likelihood": profession["night_work_likelihood"],
            "job_zone": profession["job_zone"],
            "interest_tags": sorted(set(degree["interest_tags"]) | set(profession["interest_tags"])),
            "interest_overlap": interest_matches,
            "fit_scores": degree["fit_scores"],
            "top_skills": top_skills,
            "top_work_styles": profession["top_work_styles"],
            "technology_examples": profession["technology_examples"],
            "required_education_distribution": profession["required_education_distribution"],
            "related_experience_distribution": profession["related_experience_distribution"],
            "on_job_training_distribution": profession["on_job_training_distribution"],
            "related_roles": profession["related_roles"],
            "tradeoff_summary": build_tradeoff_summary(profession, default_target_income),
            "reality_check": build_reality_check(degree, profession, path),
            "path_note": path["path_note"],
            "degree_note": degree["notes"],
            "profession_note": profession["notes"],
            "grade_proxy_note": degree["grade_proxy_note"],
            "field_of_degree_url": degree["field_of_degree_url"],
            "ooh_url": profession["ooh_url"],
            "onet_url": profession["onet_url"],
            "source_links": relation_sources,
            "source_ids": source_ids,
            "institution_examples": [],
        }
        path_outcomes.append(path_record)

    skills_out = sorted(unique_skills.values(), key=lambda row: row["skill_name"])
    interest_out = [INTEREST_META[key] for key in INTEREST_META]

    institution_programs_by_degree: dict[str, list[dict[str, object]]] = defaultdict(list)
    institution_program_count = 0
    for program in education_programs:
        degree_id = program["degree_id"]
        scorecard_source_id = f"source:scorecard:school:{program['institution_id']}"
        source_registry[scorecard_source_id] = source_link(
            scorecard_source_id,
            f"{program['institution_name']} on College Scorecard",
            program["scorecard_school_url"],
            "institution_profile",
            "Program example derived from College Scorecard institution and program-level data.",
            "U.S. Department of Education",
        )
        institution_site_source_id = f"source:school-site:{program['institution_id']}"
        if program.get("institution_url"):
            source_registry[institution_site_source_id] = source_link(
                institution_site_source_id,
                f"{program['institution_name']} website",
                program["institution_url"],
                "institution_website",
                "School website linked for direct program discovery after the data-backed match.",
                program["institution_name"],
            )

        source_ids = ["source:scorecard:dataset", scorecard_source_id]
        if program.get("institution_url"):
            source_ids.append(institution_site_source_id)

        program_record = {
            "program_option_id": program["program_option_id"],
            "degree_id": degree_id,
            "institution_id": to_int(program["institution_id"]),
            "institution_name": program["institution_name"],
            "institution_state": program["institution_state"],
            "institution_city": program["institution_city"],
            "institution_ownership": program["institution_ownership"],
            "segment_label": program["segment_label"],
            "program_title": program["program_title"],
            "program_credential_title": program["program_credential_title"],
            "program_awards_count": to_int(program["program_awards_count"]),
            "program_earnings_1yr": to_int(program["program_earnings_1yr"]),
            "program_earnings_4yr": to_int(program["program_earnings_4yr"]),
            "institution_earnings_10yr": to_int(program["institution_earnings_10yr"]),
            "completion_rate": round(to_float(program["completion_rate"]), 4),
            "admission_rate": round(to_float(program["admission_rate"]), 4),
            "sat_average": to_int(program["sat_average"]),
            "avg_net_price": to_int(program["avg_net_price"]),
            "tuition_in_state": to_int(program["tuition_in_state"]),
            "tuition_out_of_state": to_int(program["tuition_out_of_state"]),
            "featured_score": round(to_float(program["featured_score"]), 4),
            "scorecard_school_url": program["scorecard_school_url"],
            "institution_url": program["institution_url"],
            "source_links": [source_registry[source_id] for source_id in source_ids],
        }
        institution_program_count += 1
        institution_programs_by_degree[degree_id].append(program_record)

    for degree_id in institution_programs_by_degree:
        institution_programs_by_degree[degree_id] = sorted(
            institution_programs_by_degree[degree_id],
            key=lambda item: (
                item["featured_score"],
                item["program_earnings_4yr"],
                item["institution_earnings_10yr"],
                item["completion_rate"],
            ),
            reverse=True,
        )

    graph_nodes = []
    graph_links = []
    seen_nodes = set()

    for degree in degrees_out:
        node = {
            "id": degree["degree_id"],
            "label": degree["degree_name"],
            "type": "degree",
            "group": degree["degree_family"],
        }
        graph_nodes.append(node)
        seen_nodes.add(node["id"])

    for profession in professions_out:
        node = {
            "id": profession["profession_id"],
            "label": profession["profession_name"],
            "type": "profession",
            "group": profession["profession_family"],
        }
        graph_nodes.append(node)
        seen_nodes.add(node["id"])

    top_skill_ids_by_profession: dict[str, list[str]] = defaultdict(list)
    for row in profession_skill_rows:
        if len(top_skill_ids_by_profession[row["profession_id"]]) < 3:
            top_skill_ids_by_profession[row["profession_id"]].append(row["skill_id"])

    skills_lookup = {row["skill_id"]: row for row in skills_out}
    for skill_id in sorted({item for values in top_skill_ids_by_profession.values() for item in values}):
        skill = skills_lookup[skill_id]
        if skill_id not in seen_nodes:
            graph_nodes.append(
                {
                    "id": skill_id,
                    "label": skill["skill_name"],
                    "type": "skill",
                    "group": skill["skill_category"],
                }
            )
            seen_nodes.add(skill_id)

    for row in degree_to_profession_out:
        graph_links.append(
            {
                "source": row["degree_id"],
                "target": row["profession_id"],
                "type": "leads_to",
            }
        )
    for profession_id, skill_ids in top_skill_ids_by_profession.items():
        for skill_id in skill_ids:
            graph_links.append(
                {
                    "source": profession_id,
                    "target": skill_id,
                    "type": "requires",
                }
            )

    degree_options = []
    paths_by_degree: dict[str, list[dict[str, object]]] = defaultdict(list)
    for path in path_outcomes:
        paths_by_degree[path["degree_id"]].append(path)

    for degree in degrees_out:
        linked_paths = sorted(
            paths_by_degree.get(degree["degree_id"], []),
            key=lambda item: (item["confidence"], item["median_hourly_wage_usd"], item["openings_annual"]),
            reverse=True,
        )
        top_paths = []
        for path in linked_paths[:4]:
            top_paths.append(
                {
                    "path_id": path["path_id"],
                    "profession_id": path["profession_id"],
                    "profession_name": path["profession_name"],
                    "median_hourly_wage_usd": path["median_hourly_wage_usd"],
                    "median_annual_wage_usd": path["median_annual_wage_usd"],
                    "avg_weekly_hours_estimate": path["avg_weekly_hours_estimate"],
                    "schedule_type": path["schedule_type"],
                    "ai_exposure": path["ai_exposure"],
                    "geographic_concentration": path["geographic_concentration"],
                    "confidence": path["confidence"],
                    "tradeoff_summary": path["tradeoff_summary"],
                }
            )
        degree_options.append(
            {
                **degree,
                "path_count": len(linked_paths),
                "institution_examples": institution_programs_by_degree.get(degree["degree_id"], [])[:12],
                "institution_example_count": len(institution_programs_by_degree.get(degree["degree_id"], [])),
                "program_states_covered": len(
                    {
                        item["institution_state"]
                        for item in institution_programs_by_degree.get(degree["degree_id"], [])
                    }
                ),
                "top_paths": top_paths,
                "max_downstream_hourly_wage_usd": max(
                    (item["median_hourly_wage_usd"] for item in linked_paths),
                    default=0,
                ),
                "max_downstream_annual_wage_usd": max(
                    (item["median_annual_wage_usd"] for item in linked_paths),
                    default=0,
                ),
                "best_openings_annual": max((item["openings_annual"] for item in linked_paths), default=0),
            }
        )

    for row in path_outcomes:
        row["institution_examples"] = institution_programs_by_degree.get(row["degree_id"], [])[:8]

    app_payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "prototype_scope": {
            "degree_count": len(degrees_out),
            "profession_count": len(professions_out),
            "path_count": len(path_outcomes),
            "skill_count": len(skills_out),
            "program_option_count": institution_program_count,
        },
        "notes": [
            "This prototype is student-facing by design: it mixes official labor datasets with a clear lifestyle overlay.",
            "Tradeoffs like hours, hub dependence, AI exposure, and boundary quality are meant to make the life implications explicit.",
            "Where public datasets do not express a student-facing concept directly, the prototype uses conservative manual judgment and labels it as an overlay.",
        ],
        "profile_schema": {
            "journeys": [
                {
                    "journey_id": "lifestyle_first",
                    "label": "Start with the life you want",
                    "description": "Choose the lifestyle and work pattern you want, then see the education routes that make it plausible.",
                },
                {
                    "journey_id": "education_first",
                    "label": "Start with what fits you",
                    "description": "Begin from your likely fit, grades, cost tolerance, and strengths, then explore the downstream careers.",
                },
            ],
            "academic_profiles": [
                {"value": key, "label": title_case(key), "rank": rank}
                for key, rank in ACADEMIC_PROFILE_ORDER.items()
            ],
            "budget_levels": [
                {"value": key, "label": title_case(key), "rank": rank}
                for key, rank in COST_BAND_ORDER.items()
            ],
            "training_limits_years": [
                {"value": 1, "label": "Up to 1 year"},
                {"value": 2, "label": "Up to 2 years"},
                {"value": 4, "label": "Up to 4 years"},
                {"value": 6, "label": "Up to 6 years"},
                {"value": 10, "label": "Up to 10 years"},
            ],
            "strengths": [
                {"strength_id": "analytical", "label": "Analytical"},
                {"strength_id": "creative", "label": "Creative"},
                {"strength_id": "people", "label": "People-focused"},
                {"strength_id": "hands_on", "label": "Hands-on"},
            ],
        },
        "sources": sorted(source_registry.values(), key=lambda item: item["source_id"]),
        "lifestyle_tiers": lifestyle_tiers,
        "city_cost_tiers": city_cost_tiers,
        "interests": interest_out,
        "degrees": degree_options,
        "professions": professions_out,
        "paths": path_outcomes,
        "graph": {"nodes": graph_nodes, "links": graph_links},
        "graph_export_file": "./data/grc20-compatible-graph.json",
    }

    write_csv(
        DATA_DIR / "degrees_v1.csv",
        [
            {
                "degree_id": row["degree_id"],
                "degree_node_id": row["degree_node_id"],
                "degree_name": row["degree_name"],
                "degree_family": row["degree_family"],
                "cip_code": row["cip_code"],
                "award_level": row["award_level"],
                "pathway_type": row["pathway_type"],
                "source_type": row["source_type"],
                "academic_profile_min": row["academic_profile_min"],
                "academic_profile_min_rank": row["academic_profile_min_rank"],
                "time_to_complete_years": row["time_to_complete_years"],
                "cost_band": row["cost_band"],
                "cost_band_rank": row["cost_band_rank"],
                "analytical_fit": row["fit_scores"]["analytical"],
                "creative_fit": row["fit_scores"]["creative"],
                "people_fit": row["fit_scores"]["people"],
                "hands_on_fit": row["fit_scores"]["hands_on"],
                "field_of_degree_url": row["field_of_degree_url"],
                "field_employment_2023": row["field_employment_2023"],
                "field_median_annual_wage_usd": row["field_median_annual_wage_usd"],
                "field_pct_bachelors_jobs": row["field_pct_bachelors_jobs"],
                "field_pct_advanced_degree": row["field_pct_advanced_degree"],
                "field_source": row["field_source"],
                "notes": row["notes"],
                "grade_proxy_note": row["grade_proxy_note"],
                "interest_tags": ";".join(row["interest_tags"]),
                "source_ids_json": json_compact([item["source_id"] for item in row["source_links"]]),
            }
            for row in degrees_out
        ],
        [
            "degree_id",
            "degree_node_id",
            "degree_name",
            "degree_family",
            "cip_code",
            "award_level",
            "pathway_type",
            "source_type",
            "academic_profile_min",
            "academic_profile_min_rank",
            "time_to_complete_years",
            "cost_band",
            "cost_band_rank",
            "analytical_fit",
            "creative_fit",
            "people_fit",
            "hands_on_fit",
            "field_of_degree_url",
            "field_employment_2023",
            "field_median_annual_wage_usd",
            "field_pct_bachelors_jobs",
            "field_pct_advanced_degree",
            "field_source",
            "notes",
            "grade_proxy_note",
            "interest_tags",
            "source_ids_json",
        ],
    )

    write_csv(
        DATA_DIR / "professions_v1.csv",
        [
            {
                "profession_id": row["profession_id"],
                "profession_node_id": row["profession_node_id"],
                "profession_name": row["profession_name"],
                "soc_code": row["soc_code"],
                "profession_family": row["profession_family"],
                "median_annual_wage_usd": row["median_annual_wage_usd"],
                "median_hourly_wage_usd": row["median_hourly_wage_usd"],
                "employment_count": row["employment_count"],
                "outlook_growth_pct": row["outlook_growth_pct"],
                "openings_annual": row["openings_annual"],
                "entry_level_education": row["entry_level_education"],
                "typical_training_years_post_hs": row["typical_training_years_post_hs"],
                "job_zone": row["job_zone"],
                "avg_weekly_hours_estimate": row["avg_weekly_hours_estimate"],
                "schedule_type": row["schedule_type"],
                "boundary_quality": row["boundary_quality"],
                "predictability": row["predictability"],
                "ai_exposure": row["ai_exposure"],
                "geographic_concentration": row["geographic_concentration"],
                "physical_intensity": row["physical_intensity"],
                "remote_friendliness": row["remote_friendliness"],
                "weekend_work_likelihood": row["weekend_work_likelihood"],
                "night_work_likelihood": row["night_work_likelihood"],
                "onet_title": row["onet_title"],
                "onet_description": row["onet_description"],
                "ooh_url": row["ooh_url"],
                "onet_url": row["onet_url"],
                "source": row["source"],
                "interest_tags": ";".join(row["interest_tags"]),
                "source_ids_json": json_compact([item["source_id"] for item in row["source_links"]]),
            }
            for row in professions_out
        ],
        [
            "profession_id",
            "profession_node_id",
            "profession_name",
            "soc_code",
            "profession_family",
            "median_annual_wage_usd",
            "median_hourly_wage_usd",
            "employment_count",
            "outlook_growth_pct",
            "openings_annual",
            "entry_level_education",
            "typical_training_years_post_hs",
            "job_zone",
            "avg_weekly_hours_estimate",
            "schedule_type",
            "boundary_quality",
            "predictability",
            "ai_exposure",
            "geographic_concentration",
            "physical_intensity",
            "remote_friendliness",
            "weekend_work_likelihood",
            "night_work_likelihood",
            "onet_title",
            "onet_description",
            "ooh_url",
            "onet_url",
            "source",
            "interest_tags",
            "source_ids_json",
        ],
    )

    write_csv(
        DATA_DIR / "skills_v1.csv",
        skills_out,
        ["skill_id", "skill_node_id", "skill_name", "skill_category", "source"],
    )
    write_csv(
        DATA_DIR / "degree_to_profession_v1.csv",
        degree_to_profession_out,
        [
            "degree_id",
            "profession_id",
            "edge_type",
            "path_strength",
            "confidence",
            "source",
            "notes",
            "crosswalk_match",
            "additional_training_required",
            "estimated_total_training_years",
            "source_ids_json",
        ],
    )
    write_csv(
        DATA_DIR / "profession_to_skill_v1.csv",
        profession_skill_rows,
        ["profession_id", "skill_id", "edge_type", "importance_score", "level_score", "source"],
    )
    write_csv(
        DATA_DIR / "profession_lifestyle_profiles_v1.csv",
        [
            {
                "profession_id": row["profession_id"],
                "lifestyle_profile_node_id": row["lifestyle_profile_node_id"],
                "avg_weekly_hours_band": row["avg_weekly_hours_band"],
                "avg_weekly_hours_estimate": row["avg_weekly_hours_estimate"],
                "schedule_type": row["schedule_type"],
                "boundary_quality": row["boundary_quality"],
                "predictability": row["predictability"],
                "ai_exposure": row["ai_exposure"],
                "geographic_concentration": row["geographic_concentration"],
                "physical_intensity": row["physical_intensity"],
                "remote_friendliness": row["remote_friendliness"],
                "weekend_work_likelihood": row["weekend_work_likelihood"],
                "night_work_likelihood": row["night_work_likelihood"],
                "heuristic_flag": row["heuristic_flag"],
                "source": row["source"],
                "notes": row["notes"],
                "time_pressure_score": row["time_pressure_score"],
                "contact_with_others_score": row["contact_with_others_score"],
                "email_score": row["email_score"],
                "face_to_face_score": row["face_to_face_score"],
                "sitting_score": row["sitting_score"],
                "standing_score": row["standing_score"],
                "walking_score": row["walking_score"],
                "regular_schedule_pct": row["regular_schedule_pct"],
                "irregular_schedule_pct": row["irregular_schedule_pct"],
                "seasonal_schedule_pct": row["seasonal_schedule_pct"],
                "required_education_distribution_json": json_compact(row["required_education_distribution"]),
                "related_experience_distribution_json": json_compact(row["related_experience_distribution"]),
                "on_job_training_distribution_json": json_compact(row["on_job_training_distribution"]),
            }
            for row in lifestyle_out
        ],
        [
            "profession_id",
            "lifestyle_profile_node_id",
            "avg_weekly_hours_band",
            "avg_weekly_hours_estimate",
            "schedule_type",
            "boundary_quality",
            "predictability",
            "ai_exposure",
            "geographic_concentration",
            "physical_intensity",
            "remote_friendliness",
            "weekend_work_likelihood",
            "night_work_likelihood",
            "heuristic_flag",
            "source",
            "notes",
            "time_pressure_score",
            "contact_with_others_score",
            "email_score",
            "face_to_face_score",
            "sitting_score",
            "standing_score",
            "walking_score",
            "regular_schedule_pct",
            "irregular_schedule_pct",
            "seasonal_schedule_pct",
            "required_education_distribution_json",
            "related_experience_distribution_json",
            "on_job_training_distribution_json",
        ],
    )
    write_csv(
        DATA_DIR / "degree_interest_tags_v1.csv",
        degree_interest_tags,
        ["degree_id", "interest_id", "edge_type", "weight"],
    )
    write_csv(
        DATA_DIR / "profession_interest_tags_v1.csv",
        profession_interest_tags,
        ["profession_id", "interest_id", "edge_type", "weight"],
    )
    write_csv(
        DATA_DIR / "path_outcomes_v1.csv",
        [
            {
                "path_id": row["path_id"],
                "degree_id": row["degree_id"],
                "degree_name": row["degree_name"],
                "degree_family": row["degree_family"],
                "award_level": row["award_level"],
                "pathway_type": row["pathway_type"],
                "academic_profile_min": row["academic_profile_min"],
                "cost_band": row["cost_band"],
                "time_to_complete_years": row["time_to_complete_years"],
                "estimated_total_training_years": row["estimated_total_training_years"],
                "profession_id": row["profession_id"],
                "profession_name": row["profession_name"],
                "profession_family": row["profession_family"],
                "soc_code": row["soc_code"],
                "entry_level_education": row["entry_level_education"],
                "typical_training_years_post_hs": row["typical_training_years_post_hs"],
                "additional_training_required": row["additional_training_required"],
                "path_strength": row["path_strength"],
                "confidence": row["confidence"],
                "crosswalk_match": row["crosswalk_match"],
                "median_annual_wage_usd": row["median_annual_wage_usd"],
                "median_hourly_wage_usd": row["median_hourly_wage_usd"],
                "employment_count": row["employment_count"],
                "outlook_growth_pct": row["outlook_growth_pct"],
                "openings_annual": row["openings_annual"],
                "avg_weekly_hours_estimate": row["avg_weekly_hours_estimate"],
                "schedule_type": row["schedule_type"],
                "boundary_quality": row["boundary_quality"],
                "predictability": row["predictability"],
                "ai_exposure": row["ai_exposure"],
                "geographic_concentration": row["geographic_concentration"],
                "remote_friendliness": row["remote_friendliness"],
                "physical_intensity": row["physical_intensity"],
                "weekend_work_likelihood": row["weekend_work_likelihood"],
                "night_work_likelihood": row["night_work_likelihood"],
                "tradeoff_summary": row["tradeoff_summary"],
                "reality_check": row["reality_check"],
                "path_note": row["path_note"],
                "grade_proxy_note": row["grade_proxy_note"],
                "ooh_url": row["ooh_url"],
                "field_of_degree_url": row["field_of_degree_url"],
                "onet_url": row["onet_url"],
                "interest_overlap": ";".join(row["interest_overlap"]),
                "top_skills_json": json_compact(row["top_skills"]),
                "institution_examples_json": json_compact(row["institution_examples"]),
                "source_ids_json": json_compact(row["source_ids"]),
            }
            for row in path_outcomes
        ],
        [
            "path_id",
            "degree_id",
            "degree_name",
            "degree_family",
            "award_level",
            "pathway_type",
            "academic_profile_min",
            "cost_band",
            "time_to_complete_years",
            "estimated_total_training_years",
            "profession_id",
            "profession_name",
            "profession_family",
            "soc_code",
            "entry_level_education",
            "typical_training_years_post_hs",
            "additional_training_required",
            "path_strength",
            "confidence",
            "crosswalk_match",
            "median_annual_wage_usd",
            "median_hourly_wage_usd",
            "employment_count",
            "outlook_growth_pct",
            "openings_annual",
            "avg_weekly_hours_estimate",
            "schedule_type",
            "boundary_quality",
            "predictability",
            "ai_exposure",
            "geographic_concentration",
            "remote_friendliness",
            "physical_intensity",
            "weekend_work_likelihood",
            "night_work_likelihood",
            "tradeoff_summary",
            "reality_check",
            "path_note",
            "grade_proxy_note",
            "ooh_url",
            "field_of_degree_url",
            "onet_url",
            "interest_overlap",
            "top_skills_json",
            "institution_examples_json",
            "source_ids_json",
        ],
    )
    write_csv(
        DATA_DIR / "degree_options_v1.csv",
        [
            {
                "degree_id": row["degree_id"],
                "degree_name": row["degree_name"],
                "award_level": row["award_level"],
                "pathway_type": row["pathway_type"],
                "academic_profile_min": row["academic_profile_min"],
                "time_to_complete_years": row["time_to_complete_years"],
                "cost_band": row["cost_band"],
                "path_count": row["path_count"],
                "institution_example_count": row["institution_example_count"],
                "program_states_covered": row["program_states_covered"],
                "max_downstream_hourly_wage_usd": row["max_downstream_hourly_wage_usd"],
                "max_downstream_annual_wage_usd": row["max_downstream_annual_wage_usd"],
                "best_openings_annual": row["best_openings_annual"],
                "institution_examples_json": json_compact(row["institution_examples"]),
                "top_paths_json": json_compact(row["top_paths"]),
                "source_ids_json": json_compact([item["source_id"] for item in row["source_links"]]),
            }
            for row in degree_options
        ],
        [
            "degree_id",
            "degree_name",
            "award_level",
            "pathway_type",
            "academic_profile_min",
            "time_to_complete_years",
            "cost_band",
            "path_count",
            "institution_example_count",
            "program_states_covered",
            "max_downstream_hourly_wage_usd",
            "max_downstream_annual_wage_usd",
            "best_openings_annual",
            "institution_examples_json",
            "top_paths_json",
            "source_ids_json",
        ],
    )

    graph_export = {
        "@context": {
            "grc20": "https://schema.org/",
            "id": "@id",
            "type": "@type",
        },
        "space": {
            "id": "space:education-outcome-map",
            "name": "Education Outcome Map",
            "description": "A graph-friendly prototype that links education routes, professions, skills, and lifestyle implications for students.",
            "compatibility_note": "This export is GRC-20-friendly in structure: stable IDs, typed entities, typed relations, and attached provenance.",
        },
        "types": [
            {"id": "type:degree", "name": "DegreeOrCredential"},
            {"id": "type:profession", "name": "Profession"},
            {"id": "type:skill", "name": "Skill"},
            {"id": "type:lifestyle_profile", "name": "LifestyleProfile"},
            {"id": "type:interest", "name": "Interest"},
            {"id": "type:source", "name": "Source"},
        ],
        "entities": [],
        "relations": [],
        "sources": sorted(source_registry.values(), key=lambda item: item["source_id"]),
    }

    for degree in degrees_out:
        graph_export["entities"].append(
            {
                "id": degree["degree_node_id"],
                "type": "type:degree",
                "name": degree["degree_name"],
                "properties": {
                    "degree_id": degree["degree_id"],
                    "family": degree["degree_family"],
                    "award_level": degree["award_level"],
                    "pathway_type": degree["pathway_type"],
                    "cip_code": degree["cip_code"],
                    "academic_profile_min": degree["academic_profile_min"],
                    "time_to_complete_years": degree["time_to_complete_years"],
                    "cost_band": degree["cost_band"],
                    "fit_scores": degree["fit_scores"],
                    "interest_tags": degree["interest_tags"],
                },
                "source_ids": [item["source_id"] for item in degree["source_links"]],
            }
        )

    for profession in professions_out:
        graph_export["entities"].append(
            {
                "id": profession["profession_node_id"],
                "type": "type:profession",
                "name": profession["profession_name"],
                "properties": {
                    "profession_id": profession["profession_id"],
                    "soc_code": profession["soc_code"],
                    "family": profession["profession_family"],
                    "median_annual_wage_usd": profession["median_annual_wage_usd"],
                    "median_hourly_wage_usd": profession["median_hourly_wage_usd"],
                    "employment_count": profession["employment_count"],
                    "outlook_growth_pct": profession["outlook_growth_pct"],
                    "openings_annual": profession["openings_annual"],
                    "entry_level_education": profession["entry_level_education"],
                    "typical_training_years_post_hs": profession["typical_training_years_post_hs"],
                    "interest_tags": profession["interest_tags"],
                },
                "source_ids": [item["source_id"] for item in profession["source_links"]],
            }
        )
        graph_export["entities"].append(
            {
                "id": profession["lifestyle_profile_node_id"],
                "type": "type:lifestyle_profile",
                "name": f"{profession['profession_name']} lifestyle profile",
                "properties": {
                    "avg_weekly_hours_estimate": profession["avg_weekly_hours_estimate"],
                    "schedule_type": profession["schedule_type"],
                    "boundary_quality": profession["boundary_quality"],
                    "predictability": profession["predictability"],
                    "ai_exposure": profession["ai_exposure"],
                    "geographic_concentration": profession["geographic_concentration"],
                    "physical_intensity": profession["physical_intensity"],
                    "remote_friendliness": profession["remote_friendliness"],
                    "weekend_work_likelihood": profession["weekend_work_likelihood"],
                    "night_work_likelihood": profession["night_work_likelihood"],
                },
                "source_ids": [
                    item["source_id"]
                    for item in [source_registry["source:overlay:heuristics"], *profession["source_links"]]
                ],
            }
        )
        graph_export["relations"].append(
            {
                "id": f"rel:lifestyle:{profession['profession_id']}",
                "type": "HAS_LIFESTYLE_PROFILE",
                "from": profession["profession_node_id"],
                "to": profession["lifestyle_profile_node_id"],
                "properties": {},
                "source_ids": [item["source_id"] for item in profession["source_links"]],
            }
        )

    for skill in skills_out:
        graph_export["entities"].append(
            {
                "id": skill["skill_node_id"],
                "type": "type:skill",
                "name": skill["skill_name"],
                "properties": {
                    "skill_id": skill["skill_id"],
                    "category": skill["skill_category"],
                },
                "source_ids": ["source:onet:release"],
            }
        )

    for interest in interest_out:
        graph_export["entities"].append(
            {
                "id": f"interest:{interest['interest_id']}",
                "type": "type:interest",
                "name": interest["interest_name"],
                "properties": {
                    "interest_id": interest["interest_id"],
                    "description": interest["description"],
                },
                "source_ids": [],
            }
        )

    for row in degree_to_profession_out:
        graph_export["relations"].append(
            {
                "id": f"rel:path:{row['degree_id']}:{row['profession_id']}",
                "type": "LEADS_TO",
                "from": f"degree:{row['degree_id']}",
                "to": f"profession:{row['profession_id']}",
                "properties": {
                    "path_strength": row["path_strength"],
                    "confidence": row["confidence"],
                    "notes": row["notes"],
                    "crosswalk_match": row["crosswalk_match"],
                    "additional_training_required": row["additional_training_required"],
                    "estimated_total_training_years": row["estimated_total_training_years"],
                },
                "source_ids": json.loads(row["source_ids_json"]),
            }
        )

    for row in profession_skill_rows:
        graph_export["relations"].append(
            {
                "id": f"rel:skill:{row['profession_id']}:{row['skill_id']}",
                "type": "REQUIRES",
                "from": f"profession:{row['profession_id']}",
                "to": f"skill:{row['skill_id']}",
                "properties": {
                    "importance_score": row["importance_score"],
                    "level_score": row["level_score"],
                },
                "source_ids": ["source:onet:release"],
            }
        )

    for row in degree_interest_tags:
        graph_export["relations"].append(
            {
                "id": f"rel:degree-interest:{row['degree_id']}:{row['interest_id']}",
                "type": "TAGGED_WITH",
                "from": f"degree:{row['degree_id']}",
                "to": f"interest:{row['interest_id']}",
                "properties": {"weight": row["weight"]},
                "source_ids": [],
            }
        )

    for row in profession_interest_tags:
        graph_export["relations"].append(
            {
                "id": f"rel:profession-interest:{row['profession_id']}:{row['interest_id']}",
                "type": "TAGGED_WITH",
                "from": f"profession:{row['profession_id']}",
                "to": f"interest:{row['interest_id']}",
                "properties": {"weight": row["weight"]},
                "source_ids": [],
            }
        )

    write_json(DATA_DIR / "path_outcomes_v1.json", path_outcomes)
    write_json(DATA_DIR / "source_registry_v1.json", sorted(source_registry.values(), key=lambda item: item["source_id"]))
    write_json(DATA_DIR / "grc20-compatible-graph.json", graph_export)
    write_json(APP_GRAPH_JSON, graph_export)

    (APP_DIR / "data").mkdir(parents=True, exist_ok=True)
    APP_DATA_JS.write_text(
        "window.PROTOTYPE_DATA = " + json.dumps(app_payload, indent=2) + ";\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
